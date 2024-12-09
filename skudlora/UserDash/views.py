from django.shortcuts import render, redirect
from fuzzywuzzy import process, fuzz
from django.contrib.auth.models import User 
from django.contrib.auth.decorators import login_required, user_passes_test
from django.utils import timezone
from django.urls import reverse
from datetime import datetime, time
from django.core.paginator import Paginator
from datetime import timedelta
from django.db.models import Q, Case, When, Value, CharField, Avg, Min, Max
from django.http import HttpResponse
from post_receiver.models import DeviceData, APIKey
from UserDash.models import Notification, DeviceGroup, EventNotification
from access_control.models import DevicePermission, DeviceGroupPermission, DistrictGroup
from django.contrib import messages
from django.http import JsonResponse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Font
import json
from folium.plugins import MarkerCluster
from django.views.decorators.http import require_GET

def is_staff_or_superuser(user):
    return user.is_staff or user.is_superuser
def is_admin_or_arendator(user):
    return user.is_superuser or user.is_staff or user.groups.filter(name='Arendators').exists()

@login_required
@require_GET
def ajax_update_table(request):
    group_name = request.GET.get('group', '')
    address = request.GET.get('address', '')

    if group_name:
        devices = DeviceData.objects.filter(device_groups__group_name=group_name)
    elif address:
        devices = DeviceData.objects.filter(address__icontains=address)
    else:
        devices = DeviceData.objects.all()

    current_time = timezone.now()
    devices = devices.annotate(
        status=Case(
            When(time__lt=current_time - timedelta(hours=1), then=Value('Оффлайн')),
            default=Value('Онлайн'),
            output_field=CharField(),
        )
    )

    devices_list = list(devices.values('address', 'status', 'dev_eui'))
    return JsonResponse({'devices': devices_list})

@login_required
def index(request):
    query = request.GET.get('search', '')
    if request.user.is_superuser:
        devices = DeviceData.objects.all()
    else:
        ud = DeviceData.objects.filter(devicepermission__user=request.user)
        dgw = DeviceGroup.objects.filter(devicegrouppermission__user=request.user)
        dig = DeviceData.objects.filter(device_groups__in=dgw)
        dgp = DistrictGroup.objects.filter(user_permissions__user=request.user)
        dgi = DeviceGroup.objects.filter(districts__in=dgp)
        dii = DeviceData.objects.filter(device_groups__in=dgi)
        devices = (ud | dig | dii).distinct()

    if query:
        query_lower = query.lower()
        found = devices.filter(address__icontains=query_lower)
        if not found.exists():
            addrs = list(devices.values_list('id', 'address'))
            choices = [(a[0], a[1].lower()) for a in addrs]
            matched = process.extract(query_lower, [c[1] for c in choices], limit=50, scorer=fuzz.partial_ratio)
            best_ids = [i[0] for val, score in matched if score > 50 for i in choices if i[1] == val]
            devices = devices.filter(id__in=best_ids) if best_ids else devices.none()
        else:
            devices = found

    current_time = timezone.now()
    devices = devices.annotate(
        status=Case(
            When(time__lt=current_time - timedelta(hours=1), then=Value('Оффлайн')),
            default=Value('Онлайн'),
            output_field=CharField(),
        )
    )
    online_count = devices.filter(status='Онлайн').count()
    total_devices = devices.count()
    offline_count = total_devices - online_count

    if request.method == "POST":
        dev_eui = request.POST.get('dev_eui')
        if dev_eui:
            try:
                device = DeviceData.objects.get(dev_eui=dev_eui)
                Notification.objects.create(
                    message=f"Открытие двери по адресу {device.address} пользователем {request.user.username}",
                    user=request.user,
                    address=device.address,
                    dev_eui=dev_eui,
                    is_read=False,
                    notification_type='system'
                )
            except DeviceData.DoesNotExist:
                pass

    paginator = Paginator(devices, 10)
    page_number = request.GET.get('page', 1)
    page_devices = paginator.get_page(page_number)
    devices_with_coords = devices.exclude(latitude__isnull=True).exclude(longitude__isnull=True)
    
    # Получаем данные групп, которым принадлежат устройства с координатами.
    # Обратите внимание: раньше мы брали для групп только lat/lng и имя.
    # Если нам нужно группировать устройства, стоит передавать все устройства группы отдельно.
    # Но пока оставим как было, в groups_data одна точка на группу.
    groups_data = list(DeviceGroup.objects.filter(devices__in=devices_with_coords)
                       .distinct()
                       .values('group_name', 'latitude', 'longitude'))

    # Устройства, не входящие в группы
    devices_not_in_groups = devices_with_coords.filter(device_groups__isnull=True)

    # Преобразуем QuerySet устройств в список словарей
    devices_data = list(devices_not_in_groups.values('latitude', 'longitude', 'address', 'status', 'dev_eui'))


    return render(request, 'index.html', {
        'devices': page_devices,
        'search_query': query,
        'total_devices': total_devices,
        'online_count': online_count,
        'offline_count': offline_count,
        'groups_data': json.dumps(groups_data),
        'devices_data': json.dumps(devices_data),
    })



@login_required
def devices(request):
    query = request.GET.get('search', '')

    if request.user.is_superuser:
        devices = DeviceData.objects.all()
    else:
        # Устройства, явно привязанные к пользователю
        user_devices = DeviceData.objects.filter(devicepermission__user=request.user)
        
        # Устройства в группах, доступных пользователю
        device_groups_with_permission = DeviceGroup.objects.filter(devicegrouppermission__user=request.user)
        devices_in_groups = DeviceData.objects.filter(device_groups__in=device_groups_with_permission)

        # Устройства, относящиеся к районам, доступным пользователю
        district_groups_with_permission = DistrictGroup.objects.filter(
            user_permissions__user=request.user
        )
        device_groups_in_districts = DeviceGroup.objects.filter(
            districts__in=district_groups_with_permission
        )
        devices_in_districts = DeviceData.objects.filter(device_groups__in=device_groups_in_districts)

        # Объединяем устройства из всех источников и удаляем дубликаты
        devices = (user_devices | devices_in_groups | devices_in_districts).distinct()


    # Применяем фильтр поиска, если он задан
    if query:
        devices = devices.filter(
            Q(address__icontains=query) |
            Q(dev_eui__icontains=query) |
            Q(description__icontains=query)
        )

    # Обработка добавления устройства
    if request.method == 'POST':
        device_name = request.POST.get('device_name')
        dev_eui = request.POST.get('dev_eui')
        device_class = request.POST.get('device_class')
        api_key_id = request.POST.get('api_key')
        address = request.POST.get('address')
        description = request.POST.get('description')

        api_key = APIKey.objects.get(id=api_key_id) if api_key_id else None

        # Создание и сохранение нового устройства
        new_device = DeviceData.objects.create(
            device_name=device_name,
            dev_eui=dev_eui,
            device_class=device_class,
            api_key=api_key,
            address=address,
            description=description
        )

        # Назначаем разрешение пользователю на новое устройство
        DevicePermission.objects.create(
            user=request.user,
            device=new_device,
            can_manage=True  # Установите в зависимости от ваших требований
        )

        return redirect('devices')

    current_time = timezone.now() + timedelta(hours=3)
    for device in devices:
        if device.time < current_time - timedelta(hours=1):
            device.status = "Оффлайн"
        else:
            device.status = "Онлайн"

        # Расчет уровней сигнала и батареи
        if device.rssi is not None:
            if device.rssi <= -120:
                device.bar_width = 10
            elif device.rssi <= -105:
                device.bar_width = round(((device.rssi + 120) / 15 * 20) + 10)
            elif device.rssi <= -90:
                device.bar_width = round(((device.rssi + 105) / 15 * 20) + 30)
            elif device.rssi <= -75:
                device.bar_width = round(((device.rssi + 90) / 15 * 30) + 50)
            else:
                device.bar_width = 100  # Уровень по умолчанию для rssi > -75
        else:
            device.bar_width = 0  # Если rssi == None, ставим 0
                
        if device.lo_ra_snr is not None:
            if device.lo_ra_snr <= -10:
                device.snr_bar_width = 10
            elif device.lo_ra_snr <= 0:
                device.snr_bar_width = round(((device.lo_ra_snr + 10) / 10) * 30) + 10
            elif device.lo_ra_snr <= 5:
                device.snr_bar_width = round(((device.lo_ra_snr) / 5) * 40) + 40
            elif device.lo_ra_snr <= 10:
                device.snr_bar_width = round(((device.lo_ra_snr - 5) / 5) * 20) + 80
            else:
                device.snr_bar_width = 100  # Отличный сигнал
        else:
            device.snr_bar_width = 0  # Если SNR == None, ставим 0

        if device.battery_level is not None:
            device.battery_bar_width = round(device.battery_level)  # Преобразуем уровень батареи напрямую в ширину
        else:
            device.battery_bar_width = 0  # Если battery_level == None, ставим 0

    paginator = Paginator(devices, 50)  # 50 устройств на странице
    page_number = request.GET.get('page', 1)
    page_devices = paginator.get_page(page_number)

    api_keys = APIKey.objects.all()  # Получаем все доступные API ключи
    return render(request, 'devices.html', {
        'devices': page_devices, 
        'search_query': query, 
        'api_keys': api_keys,  # Для отображения API ключей в форме
    })



@login_required
def notification(request):
    # Получение параметров из POST или сессии
    items_per_page = request.POST.get('items_per_page', request.session.get('items_per_page', 25))
    try:
        items_per_page = int(items_per_page)
    except ValueError:
        items_per_page = 25
    request.session['items_per_page'] = items_per_page

    time_range = request.POST.get('time_range', request.session.get('time_range', 'all'))
    notification_type = request.POST.get('notification_type', request.session.get('notification_type', None))
    user_display = request.POST.get('user_display', request.session.get('user_display', 'all'))
    request.session['time_range'] = time_range
    request.session['notification_type'] = notification_type
    request.session['user_display'] = user_display

    # Получение уведомлений
    notifications_full = Notification.objects.all()

    current_time = timezone.now()
    if time_range == 'day':
        notifications_full = notifications_full.filter(timestamp__gte=current_time - timedelta(days=1))
    elif time_range == 'week':
        notifications_full = notifications_full.filter(timestamp__gte=current_time - timedelta(weeks=1))
    elif time_range == 'month':
        notifications_full = notifications_full.filter(timestamp__gte=current_time - timedelta(days=30))
    elif time_range == 'year':
        notifications_full = notifications_full.filter(timestamp__gte=current_time - timedelta(days=365))

    if notification_type:
        notifications_full = notifications_full.filter(notification_type=notification_type)

    # Фильтрация по пользователю, если выбрано не 'all'
    selected_user = None
    if user_display and user_display != 'all':
        try:
            selected_user = User.objects.get(id=user_display)
            notifications_full = notifications_full.filter(user=selected_user)
        except User.DoesNotExist:
            pass

    notifications_full = notifications_full.order_by('-timestamp')

    paginator = Paginator(notifications_full, items_per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    selected_notifications = request.session.get('selected_notifications', [])

    if request.method == "POST":
        page_number = request.POST.get('page', 1)
        selected_notifications = request.POST.getlist('selected_notifications')
        
        if 'select_all' in request.POST:
            if request.POST['select_all'] == 'true':
                selected_notifications = list(notifications_full.values_list('id', flat=True))
            elif request.POST['select_all'] == 'false':
                selected_notifications = []

        if 'delete_selected' in request.POST:
            Notification.objects.filter(id__in=selected_notifications).delete()
            selected_notifications = []
            request.session['selected_notifications'] = selected_notifications
            url = reverse('notification') + f'?page={page_number}'
            return redirect(url)

        if 'mark_as_read' in request.POST:
            Notification.objects.filter(id__in=selected_notifications).update(is_read=True)

        if 'notification_id' in request.POST:
            notification_id = request.POST.get('notification_id')
            try:
                n = Notification.objects.get(id=notification_id)
                n.is_read = True
                n.save()
            except Notification.DoesNotExist:
                pass

        request.session['selected_notifications'] = selected_notifications
        url = reverse('notification') + f'?page={page_number}'
        return redirect(url)

    selected_notifications = request.session.get('selected_notifications', [])

    return render(request, 'notification.html', {
        'notifications': page_obj,
        'selected_notifications': selected_notifications,
        'notification_type': notification_type,
        'items_per_page': items_per_page,
        'time_range': time_range,
        'user_display': user_display,
        'selected_user': selected_user,
        'page_range': paginator.page_range,
    })

@login_required
def user_search(request):
    if not request.user.is_authenticated:
        return JsonResponse({'results': []})
    query = request.GET.get('q', '')
    page = int(request.GET.get('page', 1))
    per_page = 30
    if request.user.is_superuser:
        base_qs = Notification.objects.all()
    else:
        base_qs = Notification.objects.filter(user=request.user)
    if query:
        users = User.objects.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        ).order_by('username')
        total = users.count()
        start = (page - 1) * per_page
        end = start + per_page
        users = users[start:end]
    else:
        latest_notifications = base_qs.order_by('-timestamp')
        seen = []
        for n in latest_notifications:
            if n.user_id not in seen:
                seen.append(n.user_id)
            if len(seen) == 20:
                break
        users = User.objects.filter(id__in=seen)
        total = len(users)
    results = [{'id': u.id, 'text': f"{u.username} ({u.get_full_name()})"} for u in users]
    return JsonResponse({
        'results': results,
        'pagination': {
            'more': False
        }
    })

@user_passes_test(is_staff_or_superuser, login_url='access_denied')
def groups(request):
    if request.method == "POST":
        # Получаем данные из формы
        group_name = request.POST.get("group_name")
        address = request.POST.get("address")
        latitude = request.POST.get("latitude")
        longitude = request.POST.get("longitude")
        device_ids = request.POST.getlist("devices")  # Получаем список выбранных устройств

        print("POST данные:", request.POST)  # Для отладки

        # Проверяем корректность и обязательные поля
        if group_name and address:
            try:
                latitude = float(latitude) if latitude else None
                longitude = float(longitude) if longitude else None

                # Создаём новую группу
                group = DeviceGroup.objects.create(
                    group_name=group_name,
                    address=address,
                    latitude=latitude,
                    longitude=longitude
                )
                print(f"Группа создана: {group}")  # Отладка

                # Добавляем устройства в группу
                if device_ids:
                    devices = DeviceData.objects.filter(id__in=device_ids)  # Получаем устройства по ID
                    group.devices.set(devices)  # Устанавливаем устройства для группы
                    print(f"Устройства добавлены: {device_ids}")  # Отладка

                return redirect('groups')  # Перенаправление на страницу групп
            except ValueError as e:
                print(f"Ошибка при конвертации: {e}")
                return render(request, 'groups.html', {
                    'devices': DeviceData.objects.all(),
                    'error': "Некорректное значение для широты или долготы."
                })
            except Exception as e:
                print(f"Общая ошибка: {e}")
                return render(request, 'groups.html', {
                    'devices': DeviceData.objects.all(),
                    'error': f"Ошибка при сохранении данных: {e}"
                })

    # При GET-запросе просто показываем страницу с устройствами
    devices = DeviceData.objects.all()
    return render(request, 'groups.html', {'devices': devices})

@login_required
@user_passes_test(is_admin_or_arendator)
def event(request):
    if request.method == 'POST' and 'delete_selected' in request.POST:
        delete_ids = request.POST.getlist('delete_ids')
        if delete_ids:
            EventNotification.objects.filter(id__in=delete_ids).delete()
            messages.success(request, "Выбранные уведомления успешно удалены.")
        else:
            messages.warning(request, "Не выбрано ни одного уведомления для удаления.")
        return redirect('event')

    dev_eui = request.GET.get('dev_eui', '').strip()
    address = request.GET.get('address', '').strip()
    user_filter = request.GET.get('user', '').strip()
    notification_type = request.GET.get('notification_type', '').strip()
    start_date = request.GET.get('start_date', '').strip()
    start_time = request.GET.get('start_time', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    end_time = request.GET.get('end_time', '').strip()
    items_per_page = request.GET.get('items_per_page', 25)
    page = request.GET.get('page')
    export_excel = request.GET.get('export_excel', 'false').lower() == 'true'

    notifications = EventNotification.objects.all()

    if dev_eui:
        notifications = notifications.filter(dev_eui__icontains=dev_eui)
    if address:
        notifications = notifications.filter(address__icontains=address)
    if is_admin_or_arendator(request.user) and user_filter:
        notifications = notifications.filter(user__username__icontains=user_filter)
    if notification_type:
        notifications = notifications.filter(notification_type=notification_type)
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            if start_time:
                start_datetime_obj = datetime.strptime(f"{start_date} {start_time}", '%Y-%m-%d %H:%M')
                notifications = notifications.filter(timestamp__gte=start_datetime_obj)
            else:
                # Устанавливаем время на начало дня
                start_datetime_obj = datetime.combine(start_date_obj.date(), time.min)
                notifications = notifications.filter(timestamp__gte=start_datetime_obj)
        except ValueError:
            messages.error(request, "Некорректный формат начальной даты или времени.")
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
            if end_time:
                end_datetime_obj = datetime.strptime(f"{end_date} {end_time}", '%Y-%m-%d %H:%M')
            else:
                # Устанавливаем время на конец дня
                end_datetime_obj = datetime.combine(end_date_obj.date(), time.max)
            notifications = notifications.filter(timestamp__lte=end_datetime_obj)
        except ValueError:
            messages.error(request, "Некорректный формат конечной даты или времени.")

    notifications = notifications.order_by('-timestamp')

    if export_excel:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Уведомления"

        # Заголовки столбцов в нужном порядке
        headers = ['Сообщение', 'Время']
        if is_admin_or_arendator(request.user):
            headers.append('Пользователь')
        headers.extend(['Тип уведомления', 'DevEUI'])

        ws.append(headers)

        # Определение стилей
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        gray_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        light_gray_fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")

        for idx, notification in enumerate(notifications, start=2):
            row = [
                notification.message,
                notification.timestamp.strftime('%Y-%m-%d %H:%M'),
            ]
            if is_admin_or_arendator(request.user):
                row.append(notification.user.username)
            row.extend([
                notification.get_notification_type_display(),
                notification.dev_eui,
            ])
            ws.append(row)

            # Применение границ
            for col_num in range(1, len(row) + 1):
                cell = ws.cell(row=idx, column=col_num)
                cell.border = thin_border

            # Применение фона для чередования строк
            fill = gray_fill if idx % 2 == 0 else light_gray_fill
            for col_num in range(1, len(row) + 1):
                ws.cell(row=idx, column=col_num).fill = fill

        # Применение границ и стилей к заголовкам
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.border = thin_border
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.font = Font(bold=True)

        # Автоматическая настройка ширины столбцов
        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=notifications.xlsx'
        wb.save(response)
        return response

    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(notifications, items_per_page)

    try:
        notifications_page = paginator.page(page)
    except PageNotAnInteger:
        notifications_page = paginator.page(1)
    except EmptyPage:
        notifications_page = paginator.page(paginator.num_pages)

    get_params = request.GET.copy()
    if 'page' in get_params:
        del get_params['page']
    get_params = get_params.urlencode()

    current_page = notifications_page.number
    total_pages = paginator.num_pages
    start_index = max(current_page - 2, 1)
    end_index = min(current_page + 2, total_pages)
    page_range = range(start_index, end_index + 1)

    context = {
        'notifications': notifications_page,
        'dev_eui': dev_eui,
        'address': address,
        'user_filter': user_filter,
        'notification_type': notification_type,
        'start_date': start_date,
        'start_time': start_time,
        'end_date': end_date,
        'end_time': end_time,
        'items_per_page': int(items_per_page),
        'is_admin': is_admin_or_arendator(request.user),
        'get_params': get_params,
        'page_range': page_range,
    }

    return render(request, 'event.html', context)